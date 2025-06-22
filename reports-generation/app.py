from flask import Flask, send_file, make_response, render_template_string
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import re
import os

app = Flask(__name__)

# Static data matching the screenshot fields
staticData = {
    'weekEnding': 'Date',
    'today': 'Date',
    'salesperson': 'Name',
    'location': 'Location',
    'days': [
        {'name': 'Monday', 'values': [14, 23, 4, 45, 22, 2, 100, 0, 0, 0, 210]},
        {'name': 'Tuesday', 'values': [23, 76, 10, 50, 54, 45, 80, 0, 0, 0, 338]},
        {'name': 'Wednesday', 'values': [4, 130, 11, 33, 67, 65, 400, 0, 0, 0, 710]},
        {'name': 'Thursday', 'values': [102, 40, 18, 0, 86, 82, 97, 0, 0, 0, 425]},
        {'name': 'Friday', 'values': [33, 55, 22, 49, 143, 26, 50, 0, 0, 0, 378]},
        {'name': 'Saturday', 'values': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]},
        {'name': 'Sunday', 'values': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]},
    ],
    'categories': [
        'IN SALES OFFICE', 'OUTSIDE OFFICE', 'IN OFFICE VISITS', 'OUTSIDE CALLS',
        'FILE PHONE CALLS', 'NEW ACCT. PHONE', 'GUEST ROOMS', 'FOOD & BEVERAGE',
        'MTG. ROOM RENTAL', 'OTHER*', 'TOTAL',
    ],
    'totals': [176, 324, 65, 177, 372, 220, 727, 0, 0, 0, 2061],
    'goal': [200, 400, 300, 65, 500, 300, 400, 600, 300, 300, 3365],
    'variance': [-24, -76, -235, 112, -128, -80, 327, -600, -300, -300, -1304],
    'explanation': '',
}

TEMPLATE_PATH = 'template.rtf'

# Helper: Fill RTF template with data
def fill_rtf_template(template, data):
    flat = {
        'salesperson': data['salesperson'],
        'weekEnding': data['weekEnding'],
        'today': data['today'],
        'location': data['location'],
    }
    result = template
    for key, value in flat.items():
        result = re.sub(r'\${' + re.escape(key) + r'}', str(value), result)
    for i, cat in enumerate(data['categories']):
        safe_cat = re.sub(r'[^A-Za-z0-9]', '', cat)
        result = re.sub(r'\${totals_' + safe_cat + r'}', str(data['totals'][i]), result)
        result = re.sub(r'\${goal_' + safe_cat + r'}', str(data['goal'][i]), result)
        result = re.sub(r'\${variance_' + safe_cat + r'}', str(data['variance'][i]), result)
    for day in data['days']:
        for i, cat in enumerate(data['categories']):
            safe_cat = re.sub(r'[^A-Za-z0-9]', '', cat)
            result = re.sub(r'\${day_' + day['name'] + '_' + safe_cat + r'}', str(day['values'][i]), result)
    return result

# Route: Home
@app.route('/')
def home():
    return '''<h2>Sales Activity Report Generator</h2><ul>
        <li><a href="/report/pdf">Download PDF Report</a></li>
        <li><a href="/report/xls">Download XLS Report</a></li>
        <li><a href="/report/rtf">Download PDF Report (from template)</a></li>
        <li><a href="/report/rtf/html">Preview RTF Report as HTML</a></li>
    </ul>'''

# Route: PDF (styled, landscape)
@app.route('/report/pdf')
def report_pdf():
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    c.setFont('Helvetica-Bold', 22)
    c.drawString(30, height-40, 'WEEKLY SALES ACTIVITY')
    c.setFont('Helvetica', 10)
    c.drawString(30, height-60, f"SALESPERSON  {staticData['salesperson']}")
    c.drawString(300, height-60, f"WEEK ENDING  {staticData['weekEnding']}")
    c.drawString(30, height-75, f"LOCATION  {staticData['location']}")
    c.drawString(300, height-75, f"TODAY'S DATE  {staticData['today']}")
    y = height-100
    # Table header
    c.setFont('Helvetica-Bold', 9)
    x = 30
    col_widths = [60, 65, 65, 65, 65, 65, 65, 65, 65, 65, 70]
    c.drawString(x, y, 'DAYS')
    for i, cat in enumerate(staticData['categories']):
        c.drawString(x + sum(col_widths[:i+1]), y, cat)
    y -= 20
    # Table rows
    c.setFont('Helvetica', 9)
    for day in staticData['days']:
        c.drawString(x, y, day['name'])
        for i, val in enumerate(day['values']):
            c.drawRightString(x + sum(col_widths[:i+1]) + col_widths[i+1] - 4, y, f"${val:.2f}" if val else "$0.00")
        y -= 20
    # Totals, Goal, Variance
    c.setFont('Helvetica-Bold', 9)
    c.drawString(x, y, 'Totals')
    for i, val in enumerate(staticData['totals']):
        c.drawRightString(x + sum(col_widths[:i+1]) + col_widths[i+1] - 4, y, f"${val:.2f}")
    y -= 20
    c.setFont('Helvetica', 9)
    c.drawString(x, y, 'GOAL')
    for i, val in enumerate(staticData['goal']):
        c.drawRightString(x + sum(col_widths[:i+1]) + col_widths[i+1] - 4, y, f"${val:.2f}")
    y -= 20
    c.setFont('Helvetica', 9)
    c.drawString(x, y, 'VARIANCE')
    for i, val in enumerate(staticData['variance']):
        c.drawRightString(x + sum(col_widths[:i+1]) + col_widths[i+1] - 4, y, f"${val:+.2f}")
    y -= 30
    c.setFont('Helvetica', 10)
    c.drawString(x, y, '*EXPLANATION')
    y -= 30
    c.drawString(x, y, 'Approval')
    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='sales-activity-report.pdf', mimetype='application/pdf')

# Route: XLS
@app.route('/report/xls')
def report_xls():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Report'
    ws.merge_cells('A1:K1')
    ws['A1'] = 'WEEKLY SALES ACTIVITY'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = 'SALESPERSON'
    ws['B2'] = staticData['salesperson']
    ws['D2'] = 'WEEK ENDING'
    ws['E2'] = staticData['weekEnding']
    ws['G2'] = 'LOCATION'
    ws['H2'] = staticData['location']
    ws['J2'] = "TODAY'S DATE"
    ws['K2'] = staticData['today']
    ws.append([])
    header = ['DAYS'] + staticData['categories']
    ws.append(header)
    for cell in ws[4]:
        cell.font = Font(bold=True, color='B97A2A')
        cell.fill = PatternFill('solid', fgColor='FFF3E0')
        cell.alignment = Alignment(horizontal='center')
    for day in staticData['days']:
        ws.append([day['name']] + [f"${v:.2f}" if v else "$0.00" for v in day['values']])
    ws.append(['Totals'] + [f"${v:.2f}" for v in staticData['totals']])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='21523B')
    ws.append(['GOAL'] + [f"${v:.2f}" for v in staticData['goal']])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='B97A2A')
        cell.fill = PatternFill('solid', fgColor='FFF3E0')
    ws.append(['VARIANCE'] + [f"${v:+.2f}" for v in staticData['variance']])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='21523B')
        cell.fill = PatternFill('solid', fgColor='E0E0E0')
    ws.append([])
    ws.append(['*EXPLANATION'])
    ws.append([''])
    ws.append(['Approval'])
    ws.append([''])
    for i, w in enumerate([12, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='sales-activity-report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Route: Download filled RTF report as PDF
@app.route('/report/rtf')
def report_rtf_pdf():
    if not os.path.exists(TEMPLATE_PATH):
        return 'RTF template not found.', 500
    with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f:
        template = f.read()
    filled = fill_rtf_template(template, staticData)
    # Simple RTF to plain text for demo
    plain = re.sub(r'\\par', '\n', filled)
    plain = re.sub(r'\\tab', '\t', plain)
    plain = re.sub(r'\{\\[^}]+\}', '', plain)
    plain = re.sub(r'\\[a-z]+[0-9]*', '', plain)
    plain = re.sub(r'\{\}|\}', '', plain)
    plain = re.sub(r'\n', '\n', plain)
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    textobject = c.beginText(30, 800)
    textobject.setFont('Helvetica', 12)
    for line in plain.split('\n'):
        textobject.textLine(line)
    c.drawText(textobject)
    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='sales-activity-report-from-template.pdf', mimetype='application/pdf')

# Route: Display filled RTF report as HTML
@app.route('/report/rtf/html')
def report_rtf_html():
    if not os.path.exists(TEMPLATE_PATH):
        return 'RTF template not found.', 500
    with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f:
        template = f.read()
    filled = fill_rtf_template(template, staticData)
    html = re.sub(r'\\par', '<br>', filled)
    html = re.sub(r'\\tab', '&emsp;', html)
    html = re.sub(r'\{\\[^}]+\}', '', html)
    html = re.sub(r'\\[a-z]+[0-9]*', '', html)
    html = re.sub(r'\{\}|\}', '', html)
    html = re.sub(r'\n', '', html)
    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>RTF Report Preview</title></head><body>{html}</body></html>"""

if __name__ == '__main__':
    app.run(debug=True) 